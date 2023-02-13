import os.path
from django import forms
from django.shortcuts import render, HttpResponse
from app01 import models


def up(request):
    if request.method == "GET":
        return render(request, 'up.html')
    # request.GET
    # request.POST
    file_object = request.FILES.get('fs')

    file_path = os.path.join('media', file_object.name)
    with open(file_path, mode='wb') as f:
        for chunk in file_object.chunks():
            f.write(chunk)

    return HttpResponse("test")


class UpModelForm(forms.ModelForm):
    name = forms.CharField(
        widget=forms.Textarea()
    )

    class Meta:
        model = models.Info
        fields = "__all__"


def up_form(request):
    if request.method == "GET":
        form = UpModelForm()
        return render(request, 'up_form.html', {'form': form})

    form = UpModelForm(data=request.POST, files=request.FILES)
    if not form.is_valid():
        return render(request, 'up_form.html', {'form': form})

    form.save()
    return HttpResponse("success")


def up_excel(request):
    if request.method == "GET":
        return render(request, 'up_excel.html')

    file_object = request.FILES.get('fs')
    print(file_object.name)

    # Excel格式数据 openpyxl模块
    from openpyxl import load_workbook
    wb = load_workbook(file_object)

    sheet = wb.worksheets[0]
    for row in sheet.iter_rows():
        print(row[0].value, row[1].value)
        models.Info.objects.create(name=row[0].value, avatar=row[1].value)

    return HttpResponse("test")
